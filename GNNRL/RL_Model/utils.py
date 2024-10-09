import os
import pickle
import argparse
import subprocess
import numpy as np
from gym_env.envs.get_TestBench import get_random, get_test

opt_passes_str = "annotation2metdadata forceattrs inferattrs coro-early lower-expect simplifycfg sroa early-cse callsite-splitting openmp-opt ipsccp called-value-propagation globalopt typepromotion argpromotion instcombine aggressive-instcombine always-inline inliner-wrapper wholeprogramdevirt module-inline inline rpo-function-attrs openmp-opt-cgscc speculative-execution jump-threading correlated-propagation libcalls-shrinkwrap tailcallelim reassociate constraint-elimination loop-simplify lcssa loop-instsimplify loop-simplifycfg licm loop-rotate simple-loop-unswitch loop-idiom indvars loop-deletion loop-unroll-full vector-combine mldst-motion gvn sccp bdce adce memcpyopt dse move-auto-init coro-elide coro-split coro-cleanup deadargelim  elim-avail-extern recompute-globalsaa float2int lower-constant-intrinsics chr loop-distribute inject-tli-mappings loop-vectorize infer-alignment loop-load-elim slp-vectorizer loop-unroll alignment-from-assumptions loop-sink instsimplify div-rem-pairs constmerge cg-profile rel-lookup-table-converter annotation-remarks verify"
opt_passes = tuple(opt_passes_str.split())

randoms = get_random(idx=0, pgm_num=70)
tests = get_test(idx=0, pgm_num=10)

randomset = {}
testset = {}

pass_fre = np.zeros((76))

for i, bm in enumerate(tests):
    pgm, path = bm
    testset[pgm] = path
for i, bm in enumerate(randoms):
    pgm, path = bm
    randomset[pgm] = path

def read_file(file_name):
    filename = file_name
    file_path = filename + ".pkl"
    with open(file_path, "rb") as file:
        datas = pickle.load(file)
        print(datas)
    return datas

def stat_pass_fre(datas):
    for data in datas:
        for i in datas[data]['passes']:
            pass_fre[i] = pass_fre[i] + 1
    print(pass_fre)
    
def print_data(data, path, testbench_type):
    mkdir_commd = "mkdir " + path
    os.system(mkdir_commd)
    f = open(path + "rl_result_passes.txt", "w+")
    index = input()
    if index == "Print All":
        for i in data:
            f.write(i + ":" + "\n")
            pgm = i + ".cc"
            cur_path = path
            cur_path = cur_path + i + "/"
            mkdir_commd  = "mkdir " + cur_path 
            os.system(mkdir_commd)
            if testbench_type == "random":
                cp_commd = "cp " + randomset[pgm] + pgm + " " + cur_path
                os.system(cp_commd)
            elif testbench_type == "test":
                cp_commd = "cp " + testset[pgm] + pgm + " " + cur_path
                os.system(cp_commd)
            clang_command = "clang -O0 -Xclang -disable-O0-optnone -emit-llvm -S " + cur_path + pgm + " -o " + cur_path + "top.0.bc"
            result = subprocess.run(clang_command, shell=True, capture_output=True)
            if result.returncode == 0:
                length = len(data[i]['passes'])
                if(length != 0):
                    for pass_order in range(0, length):
                        opt_command = "opt -passes="+opt_passes[data[i]['passes'][pass_order]]+" "+ cur_path + 'top.' +str(pass_order)+".bc -o "+ cur_path +'top.'+str(pass_order+1)+".bc"
                        os.system(opt_command)
                        llvm_dis_command = "llvm-dis " + cur_path + "top." + str(pass_order + 1) + ".bc" + " -o " + cur_path + "top." + str(pass_order + 1) + ".ll"
                        os.system(llvm_dis_command)
                        f.write(opt_passes[data[i]['passes'][pass_order]] + "\n")
    else :
        print(data[index]['passes'])
        print(data[index]['cycle'])
        f.write(index+":"+"\n")
        pgm = index + ".cc"
        path = path + index + "/"
        mkdir_commd  = "mkdir " + path
        os.system(mkdir_commd)
        if testbench_type == "random":
            cp_commd = "cp " + randomset[pgm] + pgm + " " + path
            os.system(cp_commd)
        elif testbench_type == "test":
            cp_commd = "cp " + testset[pgm] + pgm + " " + path
            os.system(cp_commd)
        clang_command = "clang -O0 -Xclang -disable-O0-optnone -emit-llvm -S " + path + pgm + " -o " + path + "top.0.bc"
        result = subprocess.run(clang_command, shell=True, capture_output=True)
        if result.returncode == 0:
            length = len(data[index]['passes'])
            if(length != 0):
                for pass_order in range(0, length):
                    opt_command = "opt -passes="+opt_passes[data[index]['passes'][pass_order]]+" "+ path + 'top.' +str(pass_order)+".bc -o "+ path +'top.'+str(pass_order+1)+".bc"
                    os.system(opt_command)
                    llvm_dis_command = "llvm-dis " + path + "top." + str(pass_order + 1) + ".bc" + " -o " + path + "top." + str(pass_order + 1) + ".ll"
                    os.system(llvm_dis_command)
                    f.write(opt_passes[data[index]['passes'][pass_order]] + "\n")       

def write_excel(data, rno):    
    import openpyxl
    workbook = openpyxl.load_workbook("/home/eeuser/Desktop/GRL-HLS/GNNRL/RL_Model/gym_env/envs/file_names.xlsx")
    sheet = workbook.active
    index = 0
    for i in data:
        sheet.cell(row=rno, column=index+2, value=i)
        print(i)

        sheet.cell(row=rno+1, column=index+2, value=data[i]['cycle'])
        print(data[i]['cycle'])
        index += 1
        workbook.save("/home/eeuser/Desktop/GRL-HLS/GNNRL/RL_Model/gym_env/envs/file_names.xlsx")
        
def build_dataset(data, testbench_type):
    target_passes = []
    for i in data:
        if i == "random46":
            continue
        print(i)
        passes = np.zeros((1,76), dtype=bool)
        for j in data[i]['passes']:
            passes[0][j] = 1
        target_passes.append(passes)
    with open(testbench_type+"_targets_passes.pkl", "ab") as fp:
        pickle.dump(target_passes, fp)
    
def boolean_string(s) -> bool:
    if s not in {'False', 'True'}:
        raise ValueError('Not a valid boolean string')
    return s == 'True'

if __name__ == "__main__":
    argparse = argparse.ArgumentParser(usage="python --print_data False --testbench_type random",description="Use for RL result analysis")
    argparse.add_argument("--write_excel",type=str, default="False")
    argparse.add_argument("--build_dataset", type=str, default="False")
    argparse.add_argument("--excel_row_number", type=int, default=1)
    argparse.add_argument("--print_data", type=str, default="False")
    argparse.add_argument("--testbench_type", type=str, default="random")
    args = argparse.parse_args()
    
    file_name = input()
    data = read_file(file_name)
    
    if boolean_string(args.write_excel):   
        write_excel(data, int(args.excel_row_number))
    elif boolean_string(args.build_dataset):
        build_dataset(data, args.testbench_type)
    elif boolean_string(args.print_data):
        print_data(data, os.getcwd()+"/"+args.testbench_type+"_"+str(os.getpid())+"/", args.testbench_type)

'''
0 1
1 3
2 5
3 7
4 9
5 11
6 13 
7 15 
'''
